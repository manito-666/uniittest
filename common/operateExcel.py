import os
from openpyxl import load_workbook

proDir=os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
excelPath = os.path.join(proDir, "testData","test.xlsx")

class OperateExcle:
    def __init__(self):
        self.open_excel=load_workbook(excelPath)

    def open_excel_sheet(self,sheet_name):
        return self.open_excel[sheet_name]

    #通过具体行列读取数据
    def from_cell_get_data(self,sheet_name,cell,row):
        open_sheet=self.open_excel_sheet(sheet_name)
        value=open_sheet[cell+str(row)].value
        return value
    #通过具体坐标读取数据
    def from_coord_get_data(self,sheet_name,x,y):
        open_sheet=self.open_excel_sheet(sheet_name)
        value=open_sheet.cell(x,y).value
        return value

    # 将测试结果写入excel
    def write_data(self,sheet_name, row, col, value):
        workbook1 = load_workbook(excelPath)
        sheet = workbook1[sheet_name]
        sheet.cell(row, col).value = value
        workbook1.save(excelPath)

    # 统计测试用例的行数
    def count_case(self,sheet_name,excelPath):
        workbook1=load_workbook(excelPath)
        sheet=workbook1[sheet_name]
        max_row=sheet.max_row
        return max_row

    #返回包含指定测试用例数据的testcase列表
    def read_data_list(self,sheet_name,case_id):
        sheet1=self.open_excel[sheet_name]
        testcase=[]
        for i in range(1,12):
            testcase.append(sheet1.cell(case_id+1,i).value)
        return testcase
